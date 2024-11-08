import os
import json
import argparse
from pathlib import Path
import cv2
import numpy as np
from pycocotools.cocoeval import COCOeval
from comix.utils import suppress_stdout, CustomCOCO, LABEL2ID
from tabulate import tabulate

from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook

# Constants
DATASET_MAPPING = {
    'eBDtheque': 'ebd',
    'comics': 'c100', 
    'popmanga': 'pop',
    'DCM': 'dcm'
}

REMOVE_CATEGORIES = ['link_sbsc', 'balloon', 'onomatopoeia']
HEADERS = ["Class", "Images", "Instances", "Matched", "AP.50", "AP.50-.95", "AR-10", "AR-100"]

def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description='Evaluate COCO predictions.')
    parser.add_argument('-n', '--dataset_name', type=str, 
                       choices=list(DATASET_MAPPING.keys()),
                       help='Name of the dataset')
    parser.add_argument('-s', '--split', type=str, default='val',
                       choices=['val', 'test'], help='Dataset split')
    parser.add_argument('-gt', '--ground_truth', type=str, default='data/comix.coco',
                       help='Path to ground truth COCO JSON file')
    parser.add_argument('-pd', '--predictions', type=str, default='data/predicts.coco',
                       help='Path to predictions COCO JSON file')
    parser.add_argument('-wn', '--weights_name', type=str, default="dass-",
                       help='Name of the model weights used to generate results')
    parser.add_argument('-o', '--output', type=str, default='out/CDF_detection_perclass.xlsx',
                       help='Path to save evaluation results')
    parser.add_argument('--save', type=int, default=None,
                       help='Number of images to save (showing GT and predictions)')
    parser.add_argument('--xlsx', action='store_true',
                       help='Save results to Excel file')
    parser.add_argument('--plotting', action='store_true',
                       help='Print evaluation results')
    parser.add_argument('--no_layout', action='store_true',
                       help='Print results in CSV format (only with --plotting)')
    return parser.parse_args()

def load_coco_data(args):
    """Load and prepare COCO ground truth and prediction data."""
    ground_truth_json = f"{args.ground_truth}/{DATASET_MAPPING[args.dataset_name]}-{args.split}.json"
    
    with suppress_stdout():
        remove_cats = [LABEL2ID[cat] for cat in REMOVE_CATEGORIES] if REMOVE_CATEGORIES else []
        coco_gt = CustomCOCO(ground_truth_json, remove_cats=remove_cats)

        if args.weights_name:
            pred_path = f'{args.predictions}/{args.dataset_name}/{args.weights_name}/val.json'
            coco_pred = CustomCOCO(pred_path)
        else:
            prediction_json = json.load(open(ground_truth_json))
            prediction_json['annotations'] = [{**ann, 'score': 1.0} for ann in prediction_json['annotations']]
            coco_pred = coco_gt.loadRes(prediction_json['annotations'])
            
    return coco_gt, coco_pred

def get_stats(coco_eval):
    """Get evaluation statistics."""
    stats = {
        "APi": coco_eval.stats[0] if len(coco_eval.stats) > 0 else 0,
        "AP50": coco_eval.stats[1] if len(coco_eval.stats) > 1 else 0,
        "ARi-1": coco_eval.stats[6] if len(coco_eval.stats) > 6 else 0,
        "ARi-10": coco_eval.stats[7] if len(coco_eval.stats) > 7 else 0,
        "ARi-100": coco_eval.stats[8] if len(coco_eval.stats) > 8 else 0,
    }
    return stats

def evaluate_category(cocoEval, catId, coco_gt):
    """Evaluate a single category and return its metrics."""
    cocoEval.params.catIds = [catId]
    
    try:
        with suppress_stdout():
            cocoEval.evaluate()
            cocoEval.accumulate()
            cocoEval.summarize()

        # Get category info and counts
        cat_info = coco_gt.loadCats([catId])[0]
        img_ids = coco_gt.getImgIds(catIds=[catId])
        ann_ids = coco_gt.getAnnIds(catIds=[catId])
        
        # Get evaluation stats
        stats = get_stats(cocoEval)
        
        metrics = {
            'name': cat_info['name'],
            'images': len(img_ids),
            'instances': len(ann_ids),
            'matched': len(cocoEval._paramsEval.imgIds),
            'AP.50': stats['AP50'],
            'AP.50-.95': stats['APi'],
            'AR-10': stats['ARi-10'],
            'AR-100': stats['ARi-100']
        }
        
        return metrics
        
    except Exception as e:
        print(f"[Error] Evaluation failed for category {catId}: {e}")
        return {
            'name': cat_info['name'] if 'cat_info' in locals() else f"category_{catId}",
            'images': 0,
            'instances': 0,
            'matched': 0,
            'AP.50': 0,
            'AP.50-.95': 0,
            'AR-10': 0,
            'AR-100': 0
        }

def format_result_row(result):
    """Format a single result row for display."""
    return [
        result['name'],
        result['images'],
        result['instances'],
        result['matched'],
        f"{result['AP.50']:.3f}",
        f"{result['AP.50-.95']:.3f}",
        f"{result['AR-10']:.3f}",
        f"{result['AR-100']:.3f}"
    ]

def print_results(args, results_data):
    """Print results based on specified format."""
    table_data = [format_result_row(result) for result in results_data]
    
    if args.no_layout:
        print(",".join(HEADERS))
        for row in table_data:
            print(",".join(str(x) for x in row))
    else:
        print(f"\nModel evaluation {args.weights_name} for {args.dataset_name} {args.split}")
        print(tabulate(table_data, headers=HEADERS, tablefmt='pipe'))
        print()

def save_to_excel(args, results_data):
    """Save results to Excel file."""
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    
    if os.path.exists(args.output):
        workbook = load_workbook(args.output)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        
    current_row = sheet.max_row + 2
    
    # Write headers
    sheet.cell(row=current_row, column=1, 
              value=f"Model evaluation {args.weights_name} for {args.dataset_name} {args.split}")
    current_row += 1
    
    for col, header in enumerate(HEADERS, start=1):
        sheet.cell(row=current_row, column=col, value=header)
    current_row += 1
    
    # Write data
    for result in results_data:
        row_data = format_result_row(result)
        for col, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=current_row, column=col, value=value)
            cell.font = Font(color="000000")
        current_row += 1
    
    workbook.save(args.output)

def draw_annotations(img, anns, color):
    """Draw bounding boxes and scores on image."""
    for ann in anns:
        x, y, w, h = [int(v) for v in ann['bbox']]
        cv2.rectangle(img, (x, y), (x + w, y + h), color, 2)
        
        if 'score' in ann:
            score = f"{ann['score']:.2f}"
            cv2.putText(img, score, (x, y - 5), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)
    return img

def save_visualizations(args, coco_gt, coco_pred):
    """Save visualization of ground truth and predictions."""
    if not args.save:
        return
        
    output_dir = Path(f"visualizations/{args.dataset_name}/{args.weights_name}")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    img_ids = coco_gt.getImgIds()
    selected_ids = np.random.choice(img_ids, min(args.save, len(img_ids)), replace=False)
    
    for img_id in selected_ids:
        img_info = coco_gt.loadImgs([img_id])[0]
        img_path = Path(args.ground_truth).parent / 'images' / img_info['file_name']
        img = cv2.imread(str(img_path))
        
        # Draw ground truth and predictions
        img_gt = draw_annotations(img.copy(), 
                                coco_gt.loadAnns(coco_gt.getAnnIds(imgIds=[img_id])), 
                                (0, 255, 0))
        img_pred = draw_annotations(img.copy(), 
                                  coco_pred.loadAnns(coco_pred.getAnnIds(imgIds=[img_id])), 
                                  (0, 0, 255))
        
        # Save combined image
        combined_img = np.hstack((img_gt, img_pred))
        output_path = output_dir / f"{img_info['file_name']}_comparison.jpg"
        cv2.imwrite(str(output_path), combined_img)

def main():
    args = parse_args()
    coco_gt, coco_pred = load_coco_data(args)
    
    # Initialize COCO evaluator and process results
    cocoEval = COCOeval(coco_gt, coco_pred, 'bbox')
    results = []
    
    # Evaluate each category
    for catId in coco_gt.getCatIds():
        cat_results = evaluate_category(cocoEval, catId, coco_gt)
        results.append(cat_results)
    
    # Output results based on arguments
    if args.plotting:
        print_results(args, results)
    
    if args.xlsx:
        save_to_excel(args, results)
    
    if args.save:
        save_visualizations(args, coco_gt, coco_pred)

if __name__ == "__main__":
    main()